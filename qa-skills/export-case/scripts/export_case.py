#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例导出工具：支持 FreeMind (.mm)、Markdown (.md)、Excel (.xlsx)

从 JSON 格式的用例数据生成三种格式文件。
"""

import json
import argparse
from pathlib import Path
from typing import Dict, List, Any, Optional
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom


# ---------- 数据解析 ----------

def parse_json_input(input_path: str) -> Dict[str, Any]:
    """解析 JSON 格式的用例数据"""
    with open(input_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ---------- FreeMind 导出 ----------

class FreeMindExporter:
    """FreeMind 导出器"""
    PRIORITY_COLORS = {'P0': '#FF6666', 'P1': '#FFCC66', 'P2': '#99FF99'}
    MODULE_BG_COLOR = '#CCCCFF'
    FUNCTION_BG_COLOR = '#E0E0E0'

    @staticmethod
    def escape_xml(text: str) -> str:
        if not text:
            return ""
        for char, escaped in [('&', '&amp;'), ('<', '&lt;'), ('>', '&gt;'), ('"', '&quot;'), ("'", '&apos;')]:
            text = text.replace(char, escaped)
        return text

    def __init__(self, root_name: str = "测试用例"):
        self.root_name = root_name

    def create_node(self, text: str, color: Optional[str] = None,
                    bg_color: Optional[str] = None, folded: bool = False) -> Element:
        node = Element('node')
        node.set('TEXT', self.escape_xml(text))
        if color:
            node.set('COLOR', color)
        if bg_color:
            node.set('BACKGROUND_COLOR', bg_color)
        node.set('FOLDED', 'true' if folded else 'false')
        return node

    def add_test_case_details(self, parent_node: Element, case: Dict[str, Any]):
        pri = case.get('优先级') or case.get('priority', '')
        if pri:
            SubElement(parent_node, 'node', TEXT=self.escape_xml(f"优先级：{pri}"))
        for key, en_key, label in [
            ('用例类型', 'type', '用例类型'), ('执行端', 'platform', '执行端'), ('冒烟标记', 'smoke', '冒烟标记')
        ]:
            val = case.get(key) or case.get(en_key, '')
            if val:
                SubElement(parent_node, 'node', TEXT=self.escape_xml(f"{label}：{val}"))
        for label, key, en_key in [
            ('前置条件', '前置条件', 'preconditions'),
            ('测试步骤', '测试步骤', 'steps'),
            ('期望结果', '期望结果', 'expected'),
        ]:
            arr = case.get(key) or case.get(en_key, [])
            if not arr:
                continue
            if isinstance(arr, str):
                arr = [arr]
            sub = SubElement(parent_node, 'node', TEXT=label, FOLDED='true')
            for i, item in enumerate(arr, 1):
                text = item if isinstance(item, str) and item.startswith(f'{i}.') else f"{i}. {item}"
                SubElement(sub, 'node', TEXT=self.escape_xml(text))

    def build_tree(self, data: Dict[str, Any]) -> Element:
        root = self.create_node(self.root_name, folded=False)
        for module in data.get('modules', []):
            m_name = module.get('name', '未命名模块')
            m_node = self.create_node(m_name, bg_color=self.MODULE_BG_COLOR, folded=False)
            root.append(m_node)
            for function in module.get('functions', []):
                f_name = function.get('name', '未命名功能')
                f_node = self.create_node(f_name, bg_color=self.FUNCTION_BG_COLOR, folded=False)
                m_node.append(f_node)
                for tp in function.get('test_points', []):
                    tp_name = tp.get('name', '未命名测试点')
                    tp_id = tp.get('id', '')
                    tp_text = f"{tp_id}：{tp_name}" if tp_id else tp_name
                    tp_node = self.create_node(tp_text, folded=False)
                    f_node.append(tp_node)
                    if tp.get('priority'):
                        SubElement(tp_node, 'node', TEXT=self.escape_xml(f"优先级：{tp['priority']}"))
                    if tp.get('source'):
                        SubElement(tp_node, 'node', TEXT=self.escape_xml(f"来源：{tp['source']}"))
                    for case in tp.get('cases', []):
                        title = case.get('title', '未命名用例')
                        cid = case.get('id', '')
                        case_text = f"{cid}：{title}" if cid else title
                        pri = case.get('优先级') or case.get('priority', 'P2')
                        case_color = self.PRIORITY_COLORS.get(pri)
                        case_node = self.create_node(case_text, color=case_color, folded=True)
                        tp_node.append(case_node)
                        self.add_test_case_details(case_node, case)
        return root

    def generate_xml(self, root: Element) -> str:
        map_elem = Element('map')
        map_elem.set('version', '1.0.1')
        map_elem.append(root)
        rough = tostring(map_elem, encoding='utf-8')
        reparsed = minidom.parseString(rough)
        return reparsed.toprettyxml(indent='  ', encoding='utf-8').decode('utf-8')

    def export(self, data: Dict[str, Any], output_path: str):
        root = self.build_tree(data)
        xml_content = self.generate_xml(root)
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)
        print(f"✅ FreeMind 已生成：{output_path}")


# ---------- Markdown 导出 ----------

class MarkdownExporter:
    """Markdown 导出器"""

    @staticmethod
    def _format_list(items: Any) -> str:
        """将数组格式化为带数字编号的多行文本"""
        if not items:
            return ""
        if isinstance(items, str):
            return items
        # 使用换行和数字编号
        formatted = []
        for i, item in enumerate(items, 1):
            text = str(item).strip()
            # 如果已经有数字前缀就不再添加
            if not text.startswith(f'{i}.'):
                text = f"{i}. {text}"
            formatted.append(text)
        return "\n    ".join(formatted)  # 缩进以保持 Markdown 格式

    def export(self, data: Dict[str, Any], output_path: str, root_name: str = "测试用例"):
        lines: List[str] = [f"# {root_name}", ""]
        for module in data.get('modules', []):
            m_name = module.get('name', '未命名模块')
            lines.append(f"## {m_name}")
            lines.append("")
            for function in module.get('functions', []):
                f_name = function.get('name', '未命名功能')
                lines.append(f"### {f_name}")
                lines.append("")
                for tp in function.get('test_points', []):
                    tp_name = tp.get('name', '未命名测试点')
                    tp_id = tp.get('id', '')
                    tp_text = f"{tp_id}：{tp_name}" if tp_id else tp_name
                    lines.append(f"#### {tp_text}")
                    lines.append("")
                    for case in tp.get('cases', []):
                        title = case.get('title', '未命名用例')
                        cid = case.get('id', '')
                        case_title = f"{cid}：{title}" if cid else title
                        lines.append(f"- **{case_title}**")
                        pri = case.get('优先级') or case.get('priority', '')
                        if pri:
                            lines.append(f"  - 优先级：{pri}")
                        for key, en_key, label in [
                            ('用例类型', 'type', '用例类型'), ('执行端', 'platform', '执行端'), ('冒烟标记', 'smoke', '冒烟标记')
                        ]:
                            v = case.get(key) or case.get(en_key, '')
                            if v:
                                lines.append(f"  - {label}：{v}")
                        pre = self._format_list(case.get('前置条件') or case.get('preconditions'))
                        if pre:
                            lines.append(f"  - **前置条件**：")
                            lines.append(f"    {pre}")
                        steps = self._format_list(case.get('测试步骤') or case.get('steps'))
                        if steps:
                            lines.append(f"  - **测试步骤**：")
                            lines.append(f"    {steps}")
                        exp = self._format_list(case.get('期望结果') or case.get('expected'))
                        if exp:
                            lines.append(f"  - **期望结果**：")
                            lines.append(f"    {exp}")
                        lines.append("")
                    lines.append("")
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        print(f"✅ Markdown 已生成：{output_path}")


# ---------- Excel 导出 ----------

def _excel_export(data: Dict[str, Any], output_path: str, root_name: str = "测试用例"):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("❌ 导出 Excel 需要安装 openpyxl：pip install openpyxl")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"
    headers = [
        "模块", "功能", "测试点ID", "测试点描述", "用例ID", "用例标题", "优先级",
        "用例类型", "执行端", "冒烟标记", "前置条件", "测试步骤", "期望结果"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.row_dimensions[1].font = Font(bold=True)
    row = 2
    for module in data.get('modules', []):
        m_name = module.get('name', '未命名模块')
        for function in module.get('functions', []):
            f_name = function.get('name', '未命名功能')
            for tp in function.get('test_points', []):
                tp_id = tp.get('id', '')
                tp_name = tp.get('name', '未命名测试点')
                for case in tp.get('cases', []):
                    def cell_val(key: str, en_key: str, multi: bool = False) -> str:
                        v = case.get(key) or case.get(en_key, [])
                        if multi and isinstance(v, list):
                            return "\n".join(str(x) for x in v)
                        return str(v) if v else ""

                    ws.cell(row=row, column=1, value=m_name)
                    ws.cell(row=row, column=2, value=f_name)
                    ws.cell(row=row, column=3, value=tp_id)
                    ws.cell(row=row, column=4, value=tp_name)
                    ws.cell(row=row, column=5, value=case.get('id', ''))
                    ws.cell(row=row, column=6, value=case.get('title', '未命名用例'))
                    ws.cell(row=row, column=7, value=case.get('优先级') or case.get('priority', ''))
                    ws.cell(row=row, column=8, value=case.get('用例类型') or case.get('type', ''))
                    ws.cell(row=row, column=9, value=case.get('执行端') or case.get('platform', ''))
                    ws.cell(row=row, column=10, value=case.get('冒烟标记') or case.get('smoke', ''))
                    ws.cell(row=row, column=11, value=cell_val('前置条件', 'preconditions', multi=True))
                    ws.cell(row=row, column=12, value=cell_val('测试步骤', 'steps', multi=True))
                    ws.cell(row=row, column=13, value=cell_val('期望结果', 'expected', multi=True))
                    for c in range(1, 14):
                        ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel 已生成：{output_path}")


# ---------- 主入口 ----------

def main():
    parser = argparse.ArgumentParser(
        description='将测试用例导出为 FreeMind / Markdown / Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python export_case.py --format freemind -i cases.json -o testcases.mm
  python export_case.py --format markdown -i cases.json -o testcases.md
  python export_case.py --format excel -i cases.json -o testcases.xlsx
  python export_case.py --format excel -i cases.json -o out.xlsx --root "电商系统测试用例"
        """
    )
    parser.add_argument('--format', '-f', required=True, choices=['freemind', 'markdown', 'excel'],
                        help='导出格式')
    parser.add_argument('--input', '-i', required=True, help='输入 JSON 文件')
    parser.add_argument('--output', '-o', required=True, help='输出文件路径')
    parser.add_argument('--root', '-r', default='测试用例', help='根节点/项目名称（用于 FreeMind 根节点与标题）')
    args = parser.parse_args()

    print(f"📖 读取：{args.input}")
    data = parse_json_input(args.input)

    total = sum(
        len(tp.get('cases', []))
        for m in data.get('modules', [])
        for f in m.get('functions', [])
        for tp in f.get('test_points', [])
    )
    print(f"📊 用例总数：{total}")

    if args.format == 'freemind':
        FreeMindExporter(root_name=args.root).export(data, args.output)
    elif args.format == 'markdown':
        MarkdownExporter().export(data, args.output, root_name=args.root)
    elif args.format == 'excel':
        _excel_export(data, args.output, root_name=args.root)

    print("✨ 完成")


if __name__ == '__main__':
    main()
